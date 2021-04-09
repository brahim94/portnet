# -*- encoding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) 2004-2010 Tiny SPRL (http://tiny.be). All Rights Reserved
#
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see http://www.gnu.org/licenses/.
#
##############################################################################


from openerp import models, fields, api, exceptions, _
from cStringIO import StringIO
import base64
import xlwt
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font


class account_treasury_forecast(models.Model):
    _inherit = 'account.treasury.forecast'

    @api.multi
    def button_report_excel(self):
        date_day = datetime.now().strftime('%d-%m-%Y')
        i=4
        j=1
        buf=StringIO()
        self.get_entries()
        html = self._gen_html_table()
        #############formatage des cellules###############################"
        ft= Font(name='Calibri',size=15,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        ###############################definition de fichier excel##############
        wb = Workbook(guess_types=True)
        ws = wb.active
        ws.title = "Rapport"

        ####################################remplir la feuille excel avec les données########################################
        date_day = datetime.now().strftime('%d-%m-%Y')
        txt1 = unicode('Tableau de trésorerie', "utf8")
        title1 = txt1+' '+str(date_day)
        ce=ws.cell(column=4, row=1, value=title1)
        ce.font=ft
        for tr in html:
            for td in tr :
                color=None
                if tr.get("bgcolor"):
                    color=tr.get("bgcolor")
                if td.get("bgcolor"):
                    color=td.get("bgcolor")
                elif i==4:
                    color="#F0F0F0"
                ce=ws.cell(column=j, row=i, value=td.text)
                if color:
                    color=color.replace('#','FF')
                    print "color :",color
                    fl=PatternFill(fill_type='solid',start_color=color,end_color=color)
                    ce.fill=fl
                j+=1
            i+=1
            j=1

        ###################################enrgistrement du fichier####################################
        wb.save(buf)

        fichier = "SITUATION_TREORERIE_"+str(date_day)+".xlsx"
        out=base64.encodestring(buf.getvalue())
        buf.close()
        vals={'data':out,'name_file':fichier}
        wizard_id = self.pool.get("report.excel.wizard").create(self._cr, self._uid, vals, context=self._context)
        return {
            'name':_("Rapport Excel d'Analyse de trésorerie"),
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'report.excel.wizard',
            'res_id':wizard_id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'domain': '[]',
            }

class report_excel_wizard(models.TransientModel):
    _name = 'report.excel.wizard'

    name_file = fields.Char(string="Nom Fichier")
    data = fields.Binary(string="Fichier")

report_excel_wizard()

