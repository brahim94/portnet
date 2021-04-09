# -*- encoding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) 2004-2010 Tiny SPRL (http://tiny.be). All Rights Reserved
#
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see http://www.gnu.org/licenses/.
#
##############################################################################

from openerp import models, fields, api, _, exceptions
from openerp.exceptions import Warning
from datetime import datetime
from cStringIO import StringIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
#from openpyxl.worksheet import ColumnDimension
import base64


# Déclaration des différents styles appliqués dans le tableau
ft1= Font(name='Calibri',size=15,bold=True,italic=False,underline='none',strike=False,color='FF000000')
ft2= Font(name='Calibri',size=12,bold=True,italic=False,underline='none',strike=False,color='FF000000')
al1=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)
fl1=PatternFill(fill_type='solid',start_color="FFDDDDDD",end_color="FFDDDDDD")
b1=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b2=Border(left=Side(border_style='medium',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'),)
b3=Border(left=Side(border_style='medium',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b4=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b5=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'),)

class comercial_indicator_activity_report(models.TransientModel):

    _name = "comercial.indicator.activity.report"

    fiscalyear_id = fields.Many2one(comodel_name="account.fiscalyear",string="Année Fiscale")

    @api.multi
    def print_report(self):
        date_day = datetime.now().strftime('%d-%m-%Y')
        i=4 # i = décalage des lignes par rapport à la première ligne du fichier excel
        j=1 # j = décalage des lignes par rapport à la première cellule du fichier excel
        cell_tab=5 # total cellule tableau IPF
        row_tab=2 # total ligne tableau IPF
        k=0;r=0;c=0
        header_table=['Indicateur','N','N-1','Ecart','Commentaires']
        indicateurs=['CA HT']
        buf=StringIO()
        fiscalyear_id_previous=self.env['account.fiscalyear'].search([('name','=',str(int(self.fiscalyear_id.name)-1))])
        # chargement de la configuration du rapport
        # valeurs par default
        config_report=self.env['config.table.xls'].search([('code','=','IAC')])
        complete_name="Indicateur D'activité Commerciale"
        title_feuille_report="Indicateur Acti Commerciale"
        title_report="Rapport Indicateur Acti Commerciale"
        name_report_out="Indicateur Acti Commerciale"
        color_entete_table="FFDDDDDD"
        if config_report:
            if config_report.title_report:
                title_report=config_report.title_report
            if config_report.title_feuille_report:
                title_feuille_report=config_report.title_feuille_report
            if config_report.name_report_out:
                name_report_out=config_report.name_report_out
            if config_report.row_start != 0:
                i=config_report.row_start
            if config_report.column_start != 0:
                j=config_report.column_start
            if config_report.couleur_tableau_header:
                color_entete_table=config_report.couleur_tableau_header.replace("#","FF")
        fl1=PatternFill(fill_type='solid',start_color=color_entete_table,end_color=color_entete_table)

        wb = Workbook()
        ws = wb.active
        ws.title = title_feuille_report
        # remplissage de la feuille excel avec les données
        title1 = title_report+' '+str(date_day)
        ce=ws.cell(row=1,column=j,value=title1)
        ce.font=ft1
        ce.alignment=al1
        ce1=ws.cell(row=1,column=j)
        ce1=ws.merge_cells(start_row=1,start_column=j,end_row=1,end_column=j+4)
        while r < row_tab :
            rr=r+i
            # Entête tableau
            if r==0 :
                for header in header_table:
                    ce=ws.cell(row=r+i,column=j+k,value=header)
                    ce.fill=fl1
                    ce.font=ft2
                    ce.alignment=al1
                    if k==0:
                        ws.column_dimensions[ce.column].width = 25.0
                    elif k==4:
                        ws.column_dimensions[ce.column].width = 20.0
                    k+=1
                k=0
            # CA :
            elif r==1:
                self.enter_cell_value(ws,rr,j,indicateurs[0],self.get_chiffre_affaire_ht(self.fiscalyear_id)[0],self.get_chiffre_affaire_ht(fiscalyear_id_previous)[0],comment="")
            r+=1
        r=0
        while r < row_tab :
            rr=r+i
            while c < cell_tab :
                cc=c+j
                ce=ws.cell(row=rr,column=cc)
                if r!=0 and r!=row_tab-1 and c==0:
                    ce.border=b3
                elif r!=0 and r!=row_tab-1 and c==4:
                        ce.border=b4
                else:
                    ce.border=b1
                c+=1
            r+=1
            c=0
        r=0
        # while r < row_tab :
        #     rr=r+i
        #     ce=ws.cell(row=rr,column=j)
        #     ce.border=b3
        #     ce=ws.cell(row=rr,column=j+4)
        #     ce.border=b4
        #     r+=1
        while c < cell_tab :
            cc=c+j
            ce=ws.cell(row=i,column=cc)
            ce.border=b2
            ce=ws.cell(row=i+row_tab-1,column=cc)
            ce.border=b5
            c+=1
        #enrgistrement du fichier
        wb.save(buf)

        fichier = name_report_out+"_"+str(date_day)+".xlsx"
        out=base64.encodestring(buf.getvalue())
        buf.close()
        vals={'data':out,'name_file':fichier}
        wizard_id = self.pool.get("report.excel.wizard").create(self._cr, self._uid, vals, context=self._context)
        return {
            'name':_("Rapport Excel "+complete_name),
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'report.excel.wizard',
            'res_id':wizard_id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'domain': '[]',
            }

    @api.one
    def enter_cell_value(self,ws,row,column,title,nc,np,comment=""):
        ce=ws.cell(row=row,column=column,value=title)
        ce=ws.cell( row=row,column=column+1,value=nc)
        ce=ws.cell(row=row,column=column+2,value=np)
        ce=ws.cell(row=row,column=column+3,value=((nc-np)/np))

    @api.one
    def get_valeur(self, fiscalyear_id, codes):
        account_obj = self.pool.get('account.account')
        result = {}
        val = 0.0
        period_ids = self.env['account.period'].search([('fiscalyear_id', '=', fiscalyear_id.id)])
        if not period_ids:
            raise exceptions.ValidationError(_("Aucune période trouvée pour l'année fiscale " + fiscalyear_id))
        query = "l.period_id between %s and %s" % (period_ids[0].id, period_ids[-1].id)
        # account_id=account_obj.search(self._cr, self._uid, [('code','in',code)])
        for code in codes:
            account_id = account_obj.search(self._cr, self._uid, [('code', '=', code)])
            if not account_id:
                result[int(code)] = {'balance': 0}

            res = account_obj._account_account__compute(self._cr, self._uid, account_id, ['balance'], None, None, query)
            for cle, valeur in res.items():
                result[int(code)] = res[cle]
        return result

    # FUNCTION RETOURNE LE CA DE L ANNEE FISCALE
    @api.one
    def get_chiffre_affaire_ht(self, fiscalyear_id):
        val = []
        calc = []
        val = self.get_valeur(fiscalyear_id, ['711', '712'])
        calc = val[0]
        caht = abs(calc[711]['balance']) + abs(calc[712]['balance'])
        return caht


comercial_indicator_activity_report()

