# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) 2004-2010 Tiny SPRL (<http://tiny.be>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

from openerp.osv import fields, osv
from openerp.tools.translate import _
from datetime import datetime
from openerp.exceptions import Warning
from cStringIO import StringIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
#from openpyxl.worksheet import ColumnDimension
import base64

### Déclaration des différents styles appliqués dans le tableau ###
# ft1 : font des titres
ft1= Font(name='Calibri',size=15,bold=True,italic=False,underline='none',strike=False,color='FFFFFFFF')
# ft2 : font des sous-titres
ft2= Font(name='Calibri',size=12,bold=True,italic=False,underline='none',strike=False,color='FF000000')
# ft3 : font du text
ft3= Font(name='Calibri',size=11,bold=True,italic=False,underline='none',strike=False,color='FF000000')
# al1 : alignement du text au centre
al1=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)
# fl1 : arrière plan couleur gris
fl1=PatternFill(fill_type='solid',start_color="FFDDDDDD",end_color="FFDDDDDD")
# fl1 : arrière plan couleur bleu niveau 1
fl2=PatternFill(fill_type='solid',start_color="FF65C1FF",end_color="FF65C1FF")
# fl1 : arrière plan couleur bleu niveau 2
fl3=PatternFill(fill_type='solid',start_color="FFB4E1FF",end_color="FFB4E1FF")
# b1,b2,b3,b4,b5 : bordure des cellules
b1=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b2=Border(left=Side(border_style='medium',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'),)
b3=Border(left=Side(border_style='medium',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b4=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b5=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'),)

class situation_budget_wizard(osv.osv_memory):

    _name = "situation.budget.wizard"
    _description = "Situation Budget Wizard"

    _columns = {
        'date_budget_start': fields.date('Date de début'),
        'date_budget_end': fields.date('Date de fin'),
    }
    _defaults = {
        'date_budget_start': datetime.now(),
        'date_budget_end': datetime.now(),
    }

    def print_budget(self, cr, uid, ids, context=None):

        # déclaration des vars
        i=6 # i = décalage des lignes par rapport à la première ligne du fichier excel
        j=1 # j = décalage des lignes par rapport à la première cellule du fichier excel
        cell_tab=9 # total cellule tableau SB
        row_tab=8 # total ligne tableau SB
        k=0;r=0;c=0
        buf=StringIO()
        date_day = datetime.now().strftime('%d-%m-%Y')
        wizard=self.browse(cr,uid,ids) # a traité date du wizard ! wizard.date_budget
        budget_obj=self.pool.get('crossovered.budget')
        move_store_obj = self.pool.get('move.store')
        account_obj = self.pool.get('account.account')
        aal_obj = self.pool.get('account.analytic.line')
        journal_obj=self.pool.get('account.analytic.journal')
        budgets=budget_obj.browse(cr,uid,context['active_ids'])
        header_table=['Ligne budgétaire','Montant Prévu','Montant Engagement','Montant Disponible','Montant Réel','Montant Théorique','Pourcentage','E/B','R/B']

        # chargement de la configuration du rapport
        # valeurs par default
        config_report_id=self.pool.get('config.table.xls').search(cr,uid,[('code','=','SB')])
        complete_name="Situation Budgetaire"
        title_feuille_report="Situation Budgetaire"
        title_report="Situation Budgetaire"
        name_report_out="Situation Budgetaire"
        color_entete_table="FFDDDDDD"
        if config_report_id:
            config_report=self.pool.get('config.table.xls').browse(cr,uid,config_report_id)
            if config_report.title_report:
                title_report=config_report.title_report
            if config_report.title_feuille_report:
                title_feuille_report=config_report.title_feuille_report
            if config_report.name_report_out:
                name_report_out=config_report.name_report_out
            if config_report.row_start != 0:
                i=config_report.row_start
            if config_report.column_start != 0:
                j=config_report.column_start
            if config_report.couleur_tableau_header:
                color_entete_table=config_report.couleur_tableau_header.replace("#","FF")
        fl1=PatternFill(fill_type='solid',start_color=color_entete_table,end_color=color_entete_table)

        # Déclaration de la feuille excel wb
        wb = Workbook(guess_types=True)
        ws = wb.active
        ws.title = title_feuille_report

        # remplissage de la feuille excel avec les données
        # ecrire titre principale du rapport
        title1 = title_report
        ce=ws.cell(row=1,column=j,value=title1)
        ce.font=ft1
        ce.alignment=al1
        ce.fill=fl2
        ce.border=b2
        ce1=ws.cell(row=1,column=j)
        ce1=ws.merge_cells(start_row=1,start_column=j,end_row=1,end_column=j+4)

        # boucler sur les budgets selectionnés
        for budget in budgets:
            # ecrire les informations du budget
            # nom
            name_budget="Budget :  "+budget.name
            ce=ws.cell(row=i-3,column=j,value=name_budget)
            ce.font=ft3
            ce.fill=fl3
            ce.border=b2
            # code
            code_budget="Code :  "+budget.code
            ce=ws.cell(row=i-3,column=j+1,value=code_budget)
            ce.font=ft3
            ce.fill=fl3
            ce.border=b2
            # durée
            date_budget="Durée :  "+wizard.date_budget_start+" - "+wizard.date_budget_end
            ce=ws.cell(row=i-3,column=j+2,value=date_budget)
            ce.font=ft3
            ce.fill=fl3
            ce.border=b2
            ce1=ws.cell(row=i-3,column=j+2)
            ce1=ws.merge_cells(start_row=i-3,start_column=j+2,end_row=i-3,end_column=j+3)
            #date d'impréssion)
            date_print="Imprimé :  "+date_day
            ce=ws.cell(row=i-3,column=j+4,value=date_print)
            ce.font=ft3
            ce.fill=fl3
            ce.border=b2

            # écrire entête du rapport
            k=0
            for header in header_table:
                ce=ws.cell(row=i,column=j+k,value=header)
                ce.fill=fl1
                ce.font=ft3
                ce.alignment=al1
                ce.border=b2
                ws.column_dimensions[ce.column].width = 26.0
                k+=1
            k=1
            if budget:
                for bl in budget.crossovered_budget_line:
                    # Libellé
                    ce=ws.cell(row=i+k,column=j,value=bl.name)
                    ce.border=b2
                    # Poste budgetaire
                    # ce=ws.cell(row=i+k,column=j,value=bl.general_budget_id.name)
                    # ce.border=b2
                    # Montant Prévu
                    ce=ws.cell(row=i+k,column=j+1,value=bl.planned_amount)
                    ce.border=b2
                    # Montant d'Engagement
                    acc_ids = [x.id for x in bl.general_budget_id.account_ids]
                    if not acc_ids:
                        raise osv.except_osv(_('Error!'),_("The Budget '%s' has no accounts!") % str(bl.general_budget_id.name))
                    acc_ids = account_obj._get_children_and_consol(cr, uid, acc_ids, context=context)
                    sql_string = "SELECT SUM(al.amount) " \
                                 "FROM account_analytic_line al " \
                                 "LEFT JOIN account_analytic_journal aj ON al.journal_id = aj.id " \
                                 "WHERE al.budget_line_id = %s " \
                                 "AND (al.date between to_date(%s,'yyyy-mm-dd') AND to_date(%s,'yyyy-mm-dd')) " \
                                 "AND al.general_account_id=ANY(%s) AND aj.commitment_journal is true"
                    sql_args = (bl.id, wizard.date_budget_start, wizard.date_budget_end, acc_ids)
                    cr.execute(sql_string, sql_args)
                    result = cr.fetchone()[0]
                    if not result:
                        result = 0.0
                    ce=ws.cell(row=i+k,column=j+2,value=abs(result))
                    ce.border=b2
                    # Montant Disponible : available_amount
                    available_amount=bl.planned_amount-bl.commitment_amount
                    ce=ws.cell(row=i+k,column=j+3,value=available_amount)
                    ce.border=b2
                    # Montant Réel : practical_amount
                    ## début traitement
                    histo=False
                    practical_amount=0
                    result_final=0
                    result_partial=0
                    # Vérifié si il y a des lignes d'historisé lié avec les lignes
                    for aa in bl.analytic_account_id:
                        for aal in aa.line_ids:
                            if aal.move_id:
                                move_store_id = move_store_obj.search(cr,uid,[('move_line_id','=',aal.move_id.id),('create_date','>=',wizard.date_budget_end)])
                                if move_store_id:
                                    histo=True
                    if histo==False:
                        practical_amount=bl.practical_amount
                    else:
                        # get purchase journal
                        journal_analytic_ids=journal_obj.search(cr,uid,[('type','=','purchase')])
                        journal_analytic=journal_obj.browse(cr,uid,journal_analytic_ids).id
                        # get general budget
                        acc_ids = [x.id for x in bl.general_budget_id.account_ids]
                        if not acc_ids:
                            raise osv.except_osv(_('Error!'),_("The Budget '%s' has no accounts!") % str(bl.general_budget_id.name))
                        acc_ids = account_obj._get_children_and_consol(cr, uid, acc_ids, context=context)
                        for aa in bl.analytic_account_id:
                            result_partial=0
                            aal_ids = aal_obj.search(cr,uid,[('account_id','=',aa.id)])
                            if aal_ids:
                                aals=aal_obj.browse(cr,uid,aal_ids)
                                for aal in aals:
                                    move_store_id = move_store_obj.search(cr,uid,[('move_line_id','=',aal.move_id.id),('create_date','>=',wizard.date_budget_end)])
                                    if move_store_id: # ajouter controle date wizard avec date move store
                                        move_store=move_store_obj.browse(cr,uid,move_store_id)[0]
                                        if move_store.debit:
                                            result_partial+=move_store.debit
                                    else:
                                        result_partial += aal.amount
                                        # cr.execute("SELECT amount FROM account_analytic_line WHERE id=%s AND account_id=%s AND journal_id = %s AND (date "
                                        #     "between to_date(%s,'yyyy-mm-dd') AND to_date(%s,'yyyy-mm-dd')) AND "
                                        #     "general_account_id=ANY(%s)", (aal.id,aa.id,journal_analytic, bl.date_from, bl.date_to,acc_ids,))
                                        # result = cr.fetchone()[0]
                                        #if result != None:
                                        #    result_partial += result
                                result_final+=result_partial

                    acc_ids = [x.id for x in bl.general_budget_id.account_ids]
                    if not acc_ids:
                        raise osv.except_osv(_('Error!'),
                                             _("The Budget '%s' has no accounts!") % str(bl.general_budget_id.name))
                    acc_ids = account_obj._get_children_and_consol(cr, uid, acc_ids, context=context)
                    sql_string = "SELECT SUM(al.amount) " \
                                 "FROM account_analytic_line al " \
                                 "LEFT JOIN account_analytic_journal aj ON al.journal_id = aj.id " \
                                 "WHERE al.budget_line_id = %s " \
                                 "AND (al.date between to_date(%s,'yyyy-mm-dd') AND to_date(%s,'yyyy-mm-dd')) " \
                                 "AND al.general_account_id=ANY(%s) AND aj.commitment_journal is not true"
                    sql_args = (bl.id, wizard.date_budget_start, wizard.date_budget_end, acc_ids)
                    cr.execute(sql_string, sql_args)
                    result = cr.fetchone()[0]
                    if not result:
                        result = 0.0
                    ce=ws.cell(row=i+k,column=j+4,value=abs(result))
                    ce.border=b2
                    ## Fin traitement
                    # Montant Théorique : theoritical_amount
                    from_date = datetime.strptime(budget.date_from, '%Y-%m-%d')
                    to_date = datetime.strptime(budget.date_to, '%Y-%m-%d')
                    todays = datetime.strptime(datetime.now().strftime('%Y-%m-%d'), '%Y-%m-%d')
                    theoritical_amount=(bl.planned_amount/((to_date - from_date).days+1))*((todays-from_date).days+1)
                    ce=ws.cell(row=i+k,column=j+5,value=theoritical_amount)
                    ce.border=b2
                    # Pourcentage : percentage
                    # percentage=(practical_amount/theoritical_amount)*100
                    ce=ws.cell(row=i+k,column=j+6,value=bl.percentage)
                    ce.border=b2
                    # E/B
                    if bl.planned_amount != 0:
                        e_b=(bl.commitment_amount/bl.planned_amount)*100
                    else:
                        e_b = 0
                    ce=ws.cell(row=i+k,column=j+7,value=e_b)
                    ce.border=b2
                    # R/B
                    if bl.planned_amount != 0:
                        r_b=(bl.practical_amount/bl.planned_amount)*100
                    else:
                        r_b = 0
                    ce=ws.cell(row=i+k,column=j+8,value=r_b)
                    ce.border=b2
                    k+=1
            i+=k+7

        #enrgistrement et préparation du fichier excel
        wb.save(buf)
        fichier = name_report_out+"_"+str(date_day)+".xlsx"
        out=base64.encodestring(buf.getvalue())
        buf.close()
        vals={'data':out,'name_file':fichier}
        wizard_id = self.pool.get("report.excel.wizard").create(cr,uid, vals, context=context)
        return {
            'name':_("Rapport Excel"+complete_name),
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'report.excel.wizard',
            'res_id':wizard_id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'domain': '[]',
            }


