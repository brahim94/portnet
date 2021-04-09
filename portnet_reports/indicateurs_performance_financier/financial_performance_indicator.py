# -*- encoding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) 2004-2010 Tiny SPRL (http://tiny.be). All Rights Reserved
#
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see http://www.gnu.org/licenses/.
#
##############################################################################

from openerp import models, fields, api, _, exceptions
from openerp.exceptions import Warning
from datetime import datetime
from cStringIO import StringIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
#from openpyxl.worksheet import ColumnDimension
import base64

# Déclaration des différents styles appliqués dans le tableau
ft1= Font(name='Calibri',size=15,bold=True,italic=False,underline='none',strike=False,color='FF000000')
ft2= Font(name='Calibri',size=12,bold=True,italic=False,underline='none',strike=False,color='FF000000')
al1=Alignment(horizontal='center',vertical='center',text_rotation=0,wrap_text=False,shrink_to_fit=False,indent=0)
fl1=PatternFill(fill_type='solid',start_color="FFDDDDDD",end_color="FFDDDDDD")
b1=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b2=Border(left=Side(border_style='medium',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'),)
b3=Border(left=Side(border_style='medium',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b4=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
b5=Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'),)

class financial_indicator_performance_report(models.TransientModel):

    _name = "financial.indicator.performance.report"

    fiscalyear_id = fields.Many2one(comodel_name="account.fiscalyear",string="Année Fiscale")

    @api.multi
    def print_report(self):
        date_day = datetime.now().strftime('%d-%m-%Y') # date du jour
        i=5 # i = décalage des lignes par rapport à la première ligne du fichier excel
        j=1 # j = décalage des lignes par rapport à la première cellule du fichier excel
        r=0;c=0;k=0
        cell_tab=5 # total cellule tableau IPF
        row_tab=19 # total ligne tableau IPF
        header_table=['RATIOS DE GESTION & DE RENTABILITE','N','N-1','ECART','COMMENTAIRE']
        ind=['FRAIS DE PERSONNEL/ CHIFFRE D\'AFFAIRES HT',
                     'FRAIS DE PERSONNEL/VALEUR AJOUTEE',
                     'CONSOMMATIONS/ CHIFFRE D\'AFFAIRES HT',
                     'VALEUR AJOUTEE/ CHIFFRE D\'AFFAIRES HT',
                     'RESULTAT NET/ CAPITAUX PROPRES',
                     'RATIOS DE FINANCEMENT',
                     'CAPACITE D\'AUTOFINANCEMENT/ VALEUR AJOUTEE',
                     'INVESTISSEMENTS (N)/ VALEUR AJOUTEE',
                     'EXEDENT BRUT D\'EXPLOITATION/REMB  DETEES (P+I)',
                     'RATIOS DE STRUCTURE FINANCIERE',
                     'FINANCEMENT  PERMANENT/ ACTIF IMMOBILISE NET',
                     'CAPITAUX PROPRES/ FINANCEMENT PERMANENT',
                     'CAPACITE D\'AUTOFINANCEMENT/ DETTES DE FINANCEMENT',
                     'RATIO D\'ENDETTEMENT= DETTES A LMT/FINANCEMENT PERMANENT',
                     'LES  RATIOS  D\'ACTIVITE  ET  DE  RENDEMENT',
                     'TAUX D\'EVOL. DU CHIFFRE D\'AFFAIRES',
                     'TAUX D\'EVOL. DES CHARGES D\'EXPLOITATION',
                     'TAUX D\'EVOL. DE LA VALEUR AJOUTEE',
                     ]
        buf=StringIO()
        fiscalyear_id_previous=self.env['account.fiscalyear'].search([('name','=',str(int(self.fiscalyear_id.name)-1))])
        # chargement de la configuration du rapport
        # valeurs par default
        config_report=self.env['config.table.xls'].search([('code','=','IPF')])
        complete_name="Indicateurs Per Financière"
        title_feuille_report="Indicateurs Per Financière"
        title_report="Rapport Indicateurs Per Financière"
        name_report_out="Indicateurs_Performance_Financière"
        color_entete_table="FFDDDDDD"
        if config_report:
            if config_report.title_report:
                title_report=config_report.title_report
            if config_report.title_feuille_report:
                title_feuille_report=config_report.title_feuille_report
            if config_report.name_report_out:
                name_report_out=config_report.name_report_out
            if config_report.row_start != 0:
                i=config_report.row_start
            if config_report.column_start != 0:
                j=config_report.column_start
            if config_report.couleur_tableau_header:
                color_entete_table=config_report.couleur_tableau_header.replace("#","FF")
        fl1=PatternFill(fill_type='solid',start_color=color_entete_table,end_color=color_entete_table)
        # definition de fichier excel
        wb = Workbook()
        ws = wb.active
        ws.title = title_feuille_report.decode('utf-8')
        # remplissage de la feuille excel avec les données
        title1 = title_report+' '+str(date_day)
        ce=ws.cell(row=1,column=j,value=title1)
        ce.font=ft1
        ce.alignment=al1
        ce1=ws.cell(row=1,column=j)
        ce1=ws.merge_cells(start_row=1,start_column=j,end_row=1,end_column=j+4)
        while r < row_tab :
            rr=r+i
            # Entête tableau
            if r==0 :
                for header in header_table:
                    ce=ws.cell(row=r+i,column=j+k,value=header)
                    ce.fill=fl1
                    ce.font=ft2
                    ce.alignment=al1
                    if k==0:
                        ws.column_dimensions[ce.column].width = 65.0
                    elif k==4:
                        ws.column_dimensions[ce.column].width = 20.0
                    k+=1
                k=0
            # FRAIS DE PERSONNEL / CHIFFRE D'AFFAIRES HT
            elif r==1 :
                self.enter_cell_value(ws,rr,j,ind[0],self.get_row1(self.fiscalyear_id)[0]*100,self.get_row1(fiscalyear_id_previous)[0]*100)
            # FRAIS DE PERSONNEL / VALEUR AJOUTEE
            elif r==2 :
                self.enter_cell_value(ws,rr,j,ind[1],self.get_row2(self.fiscalyear_id)[0]*100,self.get_row2(fiscalyear_id_previous)[0]*100)
            # CONSOMMATIONS / CHIFFRE D'AFFAIRES HT
            elif r == 3 :
                self.enter_cell_value(ws,rr,j,ind[2],self.get_row3(self.fiscalyear_id)[0]*100,self.get_row3(fiscalyear_id_previous)[0]*100)
            # VALEUR AJOUTEE / CHIFFRE D'AFFAIRES HT
            elif r == 4 :
                self.enter_cell_value(ws,rr,j,ind[3],self.get_row4(self.fiscalyear_id)[0*100],self.get_row4(fiscalyear_id_previous)[0]*100)
            # RESULTAT NET / CAPITAUX PROPRES
            elif r == 5 :
                self.enter_cell_value(ws,rr,j,ind[4],self.get_row5(self.fiscalyear_id)[0]*100,self.get_row5(fiscalyear_id_previous)[0]*100)
            # RATIOS DE FINANCEMENT
            elif r == 6 :
                self.enter_big_title(ws,rr,j,ind[5])
            # CAPACITE D'AUTOFINANCEMENT / VALEUR AJOUTEE
            elif r == 7 :
                self.enter_cell_value(ws,rr,j,ind[6],self.get_row7(self.fiscalyear_id)[0]*100,self.get_row7(fiscalyear_id_previous)[0])
            # INVESTISSEMENTS (N) / VALEUR AJOUTEE
            elif r == 8 :
                self.enter_cell_value(ws,rr,j,ind[7],self.get_row8(self.fiscalyear_id)[0]*100,self.get_row8(fiscalyear_id_previous)[0]*100)
            # EXEDENT BRUT D'EXPLOITATION / REMB  DETEES (P+I)
            elif r == 9 :
                self.enter_cell_value(ws,rr,j,ind[8],self.get_row9(self.fiscalyear_id)[0]*100,self.get_row9(fiscalyear_id_previous)[0]*100)
            # RATIOS DE STRUCTURE FINANCIERE
            elif r == 10 :
                self.enter_big_title(ws,rr,j,ind[9])
            # FINANCEMENT  PERMANENT / ACTIF IMMOBILISE NET
            elif r == 11 :
                self.enter_cell_value(ws,rr,j,ind[10],self.get_row11(self.fiscalyear_id)[0]*100,self.get_row11(fiscalyear_id_previous)[0]*100)
            # CAPITAUX PROPRES / FINANCEMENT PERMANENT
            elif r == 12 :
                self.enter_cell_value(ws,rr,j,ind[11],self.get_row12(self.fiscalyear_id)[0]*100,self.get_row12(fiscalyear_id_previous)[0]*100)
            # CAPACITE D'AUTOFINANCEMENT / DETTES DE FINANCEMENT
            elif r == 13 :
                self.enter_cell_value(ws,rr,j,ind[12],self.get_row13(self.fiscalyear_id)[0]*100,self.get_row13(fiscalyear_id_previous)[0]*100)
            # RATIO D'ENDETTEMENT= DETTES A LMT / FINANCEMENT PERMANENT
            elif r == 14 :
                self.enter_cell_value(ws,rr,j,ind[13],self.get_row14(self.fiscalyear_id)[0]*100,self.get_row14(fiscalyear_id_previous)[0]*100)
            # LES  RATIOS  D'ACTIVITE  ET  DE  RENDEMENT
            elif r == 15 :
                self.enter_big_title(ws,rr,j,ind[14])
            # TAUX D'EVOL. DU CHIFFRE D'AFFAIRES
            elif r == 16 :
                self.enter_rate(ws,rr,j,ind[15],self.get_row16(self.fiscalyear_id,fiscalyear_id_previous)[0]*100)
            # TAUX D'EVOL. DES CHARGES D'EXPLOITATION
            elif r == 17 :
                self.enter_rate(ws,rr,j,ind[16],self.get_row17(self.fiscalyear_id,fiscalyear_id_previous)*100)
            # TAUX D'EVOL. DE LA VALEUR AJOUTEE
            elif r == 18 :
                self.enter_rate(ws,rr,j,ind[17],self.get_row18(self.fiscalyear_id,fiscalyear_id_previous)*100)
            r+=1
        r=0
        while r < row_tab :
            rr=r+i
            while c < cell_tab :
                cc=c+j
                ce=ws.cell(row=rr,column=cc)
                if r in (0,6,10,15):
                    ce.border=b2
                elif c==0:
                    ce.border=b3
                elif c==4:
                    ce.border=b4
                elif r==18:
                    ce.border=b5
                else:
                    ce.border=b1
                c+=1
            r+=1
            c=0
        ###################################enrgistrement du fichier####################################
        wb.save(buf)

        fichier = name_report_out+"_"+str(date_day)+".xlsx"
        out=base64.encodestring(buf.getvalue())
        buf.close()
        vals={'data':out,'name_file':fichier}
        wizard_id = self.pool.get("report.excel.wizard").create(self._cr, self._uid, vals, context=self._context)
        return {
            'name':_("Rapport Excel "+complete_name),
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'report.excel.wizard',
            'res_id':wizard_id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'domain': '[]',
            }
    # fonction permet d'écrire les valeurs du tableau
    @api.one
    def enter_cell_value(self,ws,row,column,title,nc,np,comment=""):
            ce=ws.cell(row=row,column=column,value=title)
            ce=ws.cell(row=row,column=column+1,value=str(nc)+"%")
            ce=ws.cell(row=row,column=column+2,value=str(np)+'%')
            ce=ws.cell(row=row,column=column+3,value=str((nc-np))+"%")
    # fonction permet d'écrire les titres gras du tableau
    @api.one
    def enter_big_title(self,ws,row,column,value):
        ce=ws.cell(row=row,column=column,value=value)
        ce.fill=fl1
        ce.font=ft2
        ce.alignment=al1
        ce1=ws.cell(row=row,column=column+1)
        ce1.fill=fl1
        ce1=ws.merge_cells(start_row=row,start_column=column+1,end_row=row,end_column=column+4)
    # fonction permet d'écrire les taux d'évolution
    @api.one
    def enter_rate(self,ws,row,column,value,n):
        ce=ws.cell(row=row,column=column,value=value)
        ce=ws.cell(row=row,column=column+1,value=str(n)+"%")
        ce1=ws.cell(row=row,column=column+1)
        ce1.alignment=al1
        ce1=ws.merge_cells(start_row=row,start_column=column+1,end_row=row,end_column=column+3)
    # fonction permet de retourner les balances des comptes(code)
    @api.one
    def get_valeur(self,fiscalyear_id,codes):
        account_obj=self.pool.get('account.account')
        result={}
        val=0.0
        period_ids=self.env['account.period'].search([('fiscalyear_id','=',fiscalyear_id.id)])
        if not period_ids:
                raise exceptions.ValidationError(_("Aucune période trouvée pour l'année fiscale "+fiscalyear_id))
        query="l.period_id between %s and %s"%(period_ids[0].id,period_ids[-1].id)
       # account_id=account_obj.search(self._cr, self._uid, [('code','in',code)])
        for code in codes :
            account_id = account_obj.search(self._cr, self._uid, [('code', '=', code)])
            if not account_id :
                result[int(code)] ={'balance':0}

            res=account_obj._account_account__compute(self._cr,self._uid,account_id,['balance'],None,None,query)
            for cle,valeur in res.items():
                result[int(code)]=res[cle]
        return result
    # FUNCTION OF FRAIS DE PERSONNEL / CHIFFRE D'AFFAIRES HT R1
    @api.one
    def get_row1(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['617'])
        calc=val[0]
        total_inv=self.get_chiffre_affaire_ht(fiscalyear_id)
        if total_inv[0] == 0 :
            retour =0.0
        else :
            retour=(calc[617]['balance']/total_inv[0])
        return retour
    # FRAIS DE PERSONNEL / VALEUR AJOUTEE R2
    @api.one
    def get_row2(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['617'])
        calc=val[0]
        val_ajout=self.get_valeur_ajoutee(fiscalyear_id)
        if val_ajout[0] == 0 :
            retour = 0.0
        else:
            retour = (calc[617]['balance'] / val_ajout[0])
        return retour
    # FUNCTION OF CONSOMMATION / CHIFFRE D'AFFAIRES HT R3
    @api.one
    def get_row3(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['612'])
        calc=val[0]
        total_inv=self.get_chiffre_affaire_ht(fiscalyear_id)
        if total_inv[0] == 0 :
            retour = 0.0
        else:
            retour = (calc[612]['balance'] / total_inv[0])
        return retour
    # FUNCTION OF VALEUR AJOUTEE / CHIFFRE D'AFFAIRES HT R4
    @api.one
    def get_row4(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur_ajoutee(fiscalyear_id)
        total_inv=self.get_chiffre_affaire_ht(fiscalyear_id)
        if total_inv[0] == 0 :
            retour = 0.0
        else :
            retour=(val[0]/total_inv[0])
        return retour
    # RESULTAT NET/ CAPITAUX PROPRES R5
    @api.one
    def get_row5(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['11','7','6'])
        calc=val[0]
        if calc[11]['balance'] == 0.0 :
            retour = 0.0
        else :
            retour=(abs(calc[7]['balance'])-calc[6]['balance'])/calc[11]['balance']
        return retour
    # CAPACITE D'AUTOFINANCEMENT / VALEUR AJOUTEE R7
    @api.one
    def get_row7(self,fiscalyear_id):
        retour=0.0
        val_ajout=self.get_valeur_ajoutee(fiscalyear_id)
        caf=self.get_capacite_autofinancement(fiscalyear_id)
        if val_ajout[0]==0.0:
            retour=0.0
        else :
            retour=caf[0]/val_ajout[0]
        return retour
    # INVESTISSEMENTS (N) / VALEUR AJOUTEE R8
    @api.one
    def get_row8(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['617'])
        calc=val[0]
        val_ajout=self.get_valeur_ajoutee(fiscalyear_id)
        try:
            retour=(calc[617]['balance']/val_ajout[0])
        except ZeroDivisionError:
            print "erreur lors de l'exécution de la formule (INVESTISSEMENTS (N) / VALEUR AJOUTEE) ..."
        return retour
    # EXEDENT BRUT D'EXPLOITATION / REMB  DETEES (P+I) R9
    @api.one
    def get_row9(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['8171','7','6'])
        calc=val[0]
        if abs(calc[7]['balance'])-calc[6]['balance']==0:
            retour = 0.0
        else :
            retour = calc[8171]['balance'] / (abs(calc[7]['balance']) - calc[6]['balance'])
        return retour
    # FINANCEMENT  PERMANENT / ACTIF IMMOBILISE NET R11
    @api.one
    def get_row11(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['14',' 2','28','29'])
        calc=val[0]
        if (calc[2]['balance']-calc[28]['balance']-calc[29]['balance'])==0.0:
            retour = 0.0
        else :
            retour = (calc[14]['balance'] / (calc[2]['balance'] - calc[28]['balance'] - calc[29]['balance']))
        return retour
    # CAPITAUX PROPRES / FINANCEMENT PERMANENT R12
    @api.one
    def get_row12(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['14','11'])
        calc=val[0]
        if calc[14]['balance']==0.0:
            retour=0.0
        else :
            retour = (calc[11]['balance'] / calc[14]['balance'])
        return retour
    # CAPACITE D'AUTOFINANCEMENT / DETTES DE FINANCEMENT R13
    @api.one
    def get_row13(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['116'])
        calc=val[0]
        caf=self.get_capacite_autofinancement(fiscalyear_id)
        if calc[116]['balance'] == 0.0 :
            retour=0.0
        else :
            retour = (caf[0]/ calc[116]['balance'])
        return retour
    # RATIO D'ENDETTEMENT= DETTES A LMT / FINANCEMENT PERMANENT R14
    @api.one
    def get_row14(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['14','44','116'])
        calc=val[0]
        if calc[14]['balance']==0.0:
            retour= 0.0
        else:
            retour = (calc[116]['balance']+calc[44]['balance']) / calc[14]['balance']
        return retour
    # VALEUR AJOUTEE
    @api.one
    def get_valeur_ajoutee(self,fiscalyear_id):
        val=[]
        calc=[]
        retour=0.0
        val=self.get_valeur(fiscalyear_id,['711','712','713','714','611','612','613','614'])
        calc=val[0]
        retour= abs(calc[711]['balance'])+abs(calc[712]['balance'])+abs(calc[713]['balance'])+abs(calc[714]['balance'])-calc[611]['balance']-calc[612]['balance']-calc[613]['balance']-calc[614]['balance']
        return retour
    # TAUX EVOL. CHIFFRE AFFAIRES R16
    @api.one
    def get_row16(self, fiscalyear_id, fiscalyear_id_previous):
        val = []
        prev_val = []
        calc = []
        prev_calc = []
        retour = 0.0
        val = self.get_valeur(fiscalyear_id, ['711', '712'])
        prev_val = self.get_valeur(fiscalyear_id_previous, ['711', '712'])
        calc = val[0]
        prev_calc = prev_val[0]
        if prev_calc[711]['balance']+prev_calc[712]['balance']==0.0:
            retour=0.0
        else:
            retour=(abs(calc[711]['balance']+calc[712]['balance'])-abs(prev_calc[711]['balance']+prev_calc[712]['balance']))/abs(prev_calc[711]['balance']+prev_calc[712]['balance'])
        return retour
    # TAUX EVOL. CHARGES EXPLOITATION R17
    def get_row17(self,fiscalyear_id,fiscalyear_id_previous):
        valc=[]
        valp=[]
        calc=[]
        retour=0.0
        valc=self.get_valeur(fiscalyear_id,['61'])
        calc=valc[0]
        valp=self.get_valeur(fiscalyear_id_previous,['61'])
        calp=valp[0]
        if calp[61]['balance']==0.0:
            retour=0.0
        else:
            retour = (calc[61]['balance'] - calp[61]['balance']) / calp[61]['balance']
        return retour
    # TAUX EVOL. VALEUR AJOUTEE R18
    def get_row18(self,fiscalyear_id,fiscalyear_id_previous):
        valc=0.0
        valp=0.0
        retour=0.0
        valc=self.get_valeur_ajoutee(fiscalyear_id)[0]
        valp=self.get_valeur_ajoutee(fiscalyear_id_previous)[0]
        if valp==0.0:
            retour=0.0
        else:
            retour = (valc - valp) / valp
        return retour
    # CAPACITE D'AUTOFINANCEMENT
    @api.one
    def get_capacite_autofinancement(self, fiscalyear_id):
        val = []
        calc = []
        val = self.get_valeur(fiscalyear_id,
                              ['7','6', '619', '6196', '639', '6394', '6396','659','65963',
                               '719','7196','769','7394','7396','759','75963','751','651'])

        calc = val[0]
        caf = (abs(calc[7]['balance'])-calc[6]['balance']) + (calc[619]['balance']-calc[6196]['balance']) + \
              (calc[639]['balance']-calc[6394]['balance']-calc[6396]['balance']) + \
              (calc[659]['balance']-calc[65963]['balance']) - (abs(calc[719]['balance'])-abs(calc[7196]['balance'])) - \
              (abs(calc[769]['balance'])-abs(calc[7394]['balance']-calc[7396]['balance'])) - \
              (abs(calc[759]['balance'])-abs(calc[75963]['balance'])) - \
              abs(calc[751]['balance']) + calc[651]['balance']
        return caf
    # CHIFFRE AFFAIRES HT (avec paramétre true la fonction retourn le CA HT) (sans paramétre ou avec paramétre false la fonction retourn le CA TOTAL TTC)
    @api.one

    def get_chiffre_affaire_ht(self, fiscalyear_id):
        val = []
        calc = []
        val = self.get_valeur(fiscalyear_id, ['711', '712'])
        calc = val[0]
        caht = abs(calc[711]['balance']) + abs(calc[712]['balance'])
        return caht

financial_indicator_performance_report()

