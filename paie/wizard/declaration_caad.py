# -*- coding: utf-8 -*-
import exceptions
from openerp import models, fields, api, exceptions, _
from cStringIO import StringIO
import base64
import xlwt
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Style

class CAADWizard(models.TransientModel):
    _name = 'caad.wizard'

    start_date = fields.Date('Du')
    end_date = fields.Date('Au')

    def get_default_period(self):
        today = datetime.today().date()
        period =today.strftime("%m")+'/'+today.strftime("%Y")
        periods = self.env['account.period'].search([['name','=',period]])

        return periods or False
    periode = fields.Many2one('account.period','Période',default=get_default_period )

    @api.multi
    @api.onchange('periode')
    def onchange_period(self):
        if self.periode:
            self.start_date=self.periode.date_start
            self.end_date=self.periode.date_stop
        else:
            raise exceptions.Warning('veuillez choisir une période valide')

    @api.model
    def get_bulletin_ids(self, periode):
        domain = [
            ('periode', '=', periode),
            ('simulation','=',False)
            #('state','=','done')
            ]
        return self.env['hr.payslip'].search(domain)

    @api.multi
    def button_report_excel(self):
        ids=self.get_bulletin_ids(self.periode.id)
        i=4
        j=1
        buf=StringIO()
        #############formatage des cellules###############################"
        ft= Font(name='Calibri',size=15,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
        ###############################definition de fichier excel##############
        wb = Workbook(guess_types=True)
        ws = wb.active
        ws.title = unicode("Déclrations CAAD")
        ####################################remplir la feuille excel avec les données########################################
        txt1 = unicode('Déclarations des cotisations CAAD', "utf8")
        title1 = txt1+' '+str(self.periode.name)
        ce=ws.cell(column=3, row=1, value=title1)
        ce.font=ft
        ft= Font(name='Calibri',size=13,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')

        borders = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))
        ce.font=ft
        my_style = Style(border=borders)
        ce.border=borders
        ce=ws.cell(column=1, row=i, value='Matricule')
        ce.font=ft
        ce=ws.cell(column=2, row=i, value='Nom')
        ce.font=ft
        ce=ws.cell(column=3, row=i, value='Prénom')
        ce.font=ft
        ce=ws.cell(column=4, row=i, value='SBI')
        ce.font=ft
        ce=ws.cell(column=5, row=i, value='CO salariales')
        ce.font=ft
        ce=ws.cell(column=6, row=i, value='CO patronales')
        ce.font=ft
        ce=ws.cell(column=7, row=i, value='Total')
        ce.font=ft
        i=i+1
        total_emp=0
        total_sbi=0
        total_cot_sal=0
        total_cot_pat=0
        total_cot=0
        for bp in ids:
            ce=ws.cell(column=j, row=i, value=bp.employee_id.matricule)
            j+=1
            ce=ws.cell(column=j, row=i, value=bp.employee_id.nom)
            j+=1
            ce=ws.cell(column=j, row=i, value=bp.employee_id.prenom)
            j+=1
            for rule in bp.details_by_salary_rule_category:
                if rule.code=="BRUT":
                    ce=ws.cell(column=j,row=i,value=rule.total)
                    total_sbi+=rule.total
            j+=1
            for rule in bp.details_by_salary_rule_category:
                if rule.code=="caad":
                    ce=ws.cell(column=j,row=i,value=rule.total)
                    total_emp+=rule.total
                    total_cot_sal+=rule.total
            j+=1
            for rule in bp.details_by_salary_rule_category:
                if rule.code=="part_pat_caad":
                    ce=ws.cell(column=j,row=i,value=rule.total)
                    total_emp+=rule.total
                    total_cot_pat+=rule.total
            j+=1
            ce=ws.cell(column=j, row=i, value=total_emp)
            i+=1
            j=1
            total_emp=0
        ce=ws.cell(column=4, row=i, value=total_sbi)
        ce=ws.cell(column=5, row=i, value=total_cot_sal)
        ce=ws.cell(column=6, row=i, value=total_cot_pat)
        ce=ws.cell(column=7, row=i, value=total_cot_pat+total_cot_sal)
        wb.save(buf)
        fichier = "Déclaration CAAD"+str(self.periode.name)+".xlsx"
        out=base64.encodestring(buf.getvalue())
        buf.close()
        vals={'data':out,'name_file':fichier}
        wizard_id = self.pool.get('dec.caad.wizard').create(self._cr, self._uid, vals, context=self._context)
        return {
            'name':_("Rapport Excel test"),
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'dec.caad.wizard',
            'res_id':wizard_id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'domain': '[]',
            }


class DECCAADWizard(models.TransientModel):
    _name = 'dec.caad.wizard'

    name_file = fields.Char('Nom fichier')
    data = fields.Binary('Fichier')


