# -*- coding: utf-8 -*-
import locale
from openerp import models, fields, api
from openerp import exceptions

from datetime import datetime, timedelta
import time
from dateutil import relativedelta
from openerp import netsvc
from openerp.tools import (
    DEFAULT_SERVER_DATE_FORMAT,
    )
from openerp.tools import float_compare, float_is_zero
from lxml import etree
from openerp.osv.orm import setup_modifiers
import re



from cStringIO import StringIO
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
#from openpyxl.styles import Style

class HrPayslipWizard(models.TransientModel):
    _name = 'hr.payslip.wizard'

    start_date = fields.Date('Du', default=fields.Date.today,translate=True)
    end_date = fields.Date('Au', default=fields.Date.today,translate=True)


    def get_default_period(self):
        today = datetime.today().date()
        period =today.strftime("%m")+'/'+today.strftime("%Y")
        periods = self.env['account.period'].search([['name','=',period]])
        return periods or False
    periode = fields.Many2one('account.period','Période',default=get_default_period )

    @api.multi
    @api.onchange('periode')
    def onchange_period(self):
        if self.periode:
            self.start_date=self.periode.date_start
            self.end_date=self.periode.date_stop
        else:
            raise exceptions.Warning('veuillez choisir une période valide')

    @api.model
    def get_bulletin_ids(self,periode):
        domain = [
            ('periode', '=', periode),
            ('simulation','=',False)
            #('state','=','done')
            ]
        return self.env['hr.payslip'].search(domain)

    @api.multi
    def get_bulletin_done_ids(self):
        ids=self.get_bulletin_ids(self.periode.id)
        list_ids = []
        i = 0
        while i <len(ids)/3.0:
            list_ids.append(ids[3*i:(i+1)*3])
            i += 1
        if list_ids:
            return list_ids
        else:
            raise exceptions.Warning('Il n y a aucun bulletin de paie durant cette période')
    @api.multi
    def get_date_now(self):

        now = datetime.now()
        return now.strftime("%d/%m/%Y")

    @api.multi
    def get_time_now(self):
        now = datetime.now()
        return now.strftime("%H:%M")

    @api.multi
    def get_pret(self, emp):
        dict = {'paid_amount': 0.00}
        reprise_interet = 0
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        payslip = self.env['hr.payslip'].search([['employee_id', '=', emp], ['periode', '=', self.periode.id]])
        loan_ids = self.env['hr.loan.line'].search(
            [['employee_id', '=', emp], ['paid_period', '=', self.periode.id]])
        for rule in payslip[0].details_by_salary_rule_category:
            if rule.code == 'regul_ded_pret_immob':
                reprise_interet = rule.total
        if loan_ids:
            dict['paid_amount'] = loan_ids[0].loan_interest + reprise_interet
        return dict

    @api.multi
    def get_deduction_logement(self, emp):
        pret_immob = 0
        reprise_interet = 0
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        payslip = self.env['hr.payslip'].search([['employee_id', '=', emp], ['periode', '=', self.periode.id]])

        for rule in payslip[0].details_by_salary_rule_category:
            if rule.code == 'regul_ded_pret_immob':
                reprise_interet = rule.total
            if rule.code == 'ded_pret_immob':
                pret_immob = rule.total
        return pret_immob+reprise_interet

    @api.multi
    def get_retraite_comp(self, emp):
        retraite = 0
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        payslip = self.env['hr.payslip'].search([['employee_id', '=', emp], ['periode', '=', self.periode.id]])
        for rule in payslip[0].details_by_salary_rule_category:
            if rule.code == 'epargne_retraite_v':
                retraite = retraite + rule.total
            if rule.code == 'epargne_retraite':
                retraite = retraite + rule.total
        return retraite

    @api.multi
    def get_conge_pris(self, emp):
        iccp = 0
        njr = 0
        payslip = self.env['hr.payslip'].search([['employee_id', '=', emp], ['periode', '=', self.periode.id]])
        for rule in payslip[0].rubriques_ids:
            if rule.name.code == 'iccp':
                iccp = rule.montant

        for rule in payslip[0].rubriques_ids:
            if rule.name.code == 'NJR':
                njr = rule.montant

        return payslip[0].nbr_jours+iccp+njr

    @api.multi
    def get_pret_locale(self, emp):
        return locale.format("%.2f", self.get_pret(emp)['paid_amount'], grouping=True)

    @api.multi
    def get_deduction_logement_locale(self, emp):
        return locale.format("%.2f", self.get_deduction_logement(emp), grouping=True)

    @api.multi
    def get_retraite_locale(self, emp):
        return locale.format("%.2f", self.get_retraite_comp(emp), grouping=True)

    @api.multi
    def get_conge_pris_locale(self, emp):
        return locale.format("%.2f", self.get_conge_pris(emp), grouping=True)

    @api.multi
    def get_total_pret(self):
        dict = {'total_paid_amount': 0.00}
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        ids = self.get_bulletin_ids(self.periode.id)
        for pay in ids:
            dict['total_paid_amount'] = dict['total_paid_amount']+self.get_pret(pay.employee_id.id)['paid_amount']
        dict['total_paid_amount'] = locale.format("%.2f", dict['total_paid_amount'], grouping=True)
        return dict

    @api.multi
    def get_total_deduction_logement(self):
        total_deduction = 0
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        ids = self.get_bulletin_ids(self.periode.id)
        for pay in ids:
            total_deduction = total_deduction + self.get_deduction_logement(pay.employee_id.id)
        total_deduction = locale.format("%.2f", total_deduction, grouping=True)
        return total_deduction

    @api.multi
    def get_total_retraite_comp(self):
        total = 0
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        ids = self.get_bulletin_ids(self.periode.id)
        for pay in ids:
            total = total + self.get_retraite_comp(pay.employee_id.id)
        total = locale.format("%.2f", total, grouping=True)
        return total

    @api.multi
    def get_total_conge_pris(self):
        total = 0
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        ids = self.get_bulletin_ids(self.periode.id)
        for pay in ids:
            total = total + self.get_conge_pris(pay.employee_id.id)
        return locale.format("%.2f", total, grouping=True)


    @api.multi
    def get_sum(self):
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
        ids=self.get_bulletin_ids(self.periode.id)
        total={'rappel_salaire_jrs': 0.00, 'nbr_jours':0.00,'base':0.00,'congpaye2':0.00,
               'allocation':0.00,'fraisdeplacement':0.00,'reprcong2':0.00,
               'rappelsursalaire':0.00,'responsabilite':0.00,'anciennete':0.00,
               'representation':0.00,'regul_rep':0.00,'transport':0.00,
               'regul_trans':0.00,'caisse':0.00,'totbrut':0.00,'rcarplafonne':0.00,
               'rcarcomplement':0.00,'ded_logement':0.00,'cnops':0.00,'modep':0.00,
               'caad':0.00,'saham':0.00,'fpro':0.00,'totcot':0.00,'ir':0.00,'arrprec':0.00,
               'arrencours':0.00,'nbrpersoch':0.00,'dedperso':0.00,
               # 'presence_jrs':0.00+26*len(ids),'brutcnss':0.00,'cot_sal':0.00,
               'presence_jrs':0.00,'brutcnssma':0.00,'cot_sal':0.00,
               'cot_pat':0.00,'net':0.00,'net_imp':0.00,'net_impr':0.00,'cout':0.00,
               'presence':0.00+191*len(ids),'jours_acquis':0.00,
               'notes_frais':0.00,'hrs_trav':1.00+191*len(ids),'reprise':0.00,'iccp':0.00,
               'prime_naissance':0.00,'prime_astreinte':0.00,'reprisesalaire':0.00,'rappel_allocation':0.00,
               'congespayes': 0.00, 'repriseconge': 0.00, 'reprise_absence_irr': 0.00, 'stc': 0.00, 'prime_except': 0.00,
               'prime_caisse': 0.00, 'prime_merite': 0.00, 'regul_ind_caisse': 0.00, 'epargne_retraite_v': 0.00,
               'pret_immob': 0.00, 'remb': 0.00, 'indlogement': 0.00, 'rappel_indlogement': 0.00,
               'indfonction': 0.00, 'rappel_indfonction': 0.00, 'indperformance': 0.00, 'rappel_indperformance': 0.00,
               'rappel_responsabilite': 0.00, 'rappel_representation': 0.00, 'prime_adha': 0.00, 'avance': 0.00,  'recore_sal': 0.00,
               'prime_scolarite': 0.00, 'prime_achoura_impo': 0.00, 'prime_achoura': 0.00, 'reprise_abs': 0.00}
        for i in ids:
            for rule in i.details_by_salary_rule_category:
                if rule.code == 'avance':
                    total['avance'] = total['avance']+rule.total
                if rule.code == 'IndLogement':
                    total['indlogement'] = total['indlogement']+rule.total
                if rule.code == 'prime_adha':
                    total['prime_adha'] = total['prime_adha']+rule.total
                if rule.code == 'RIndLogement':
                    total['rappel_indlogement'] = total['rappel_indlogement']+rule.total
                if rule.code == 'IndFonction':
                    total['indfonction'] = total['indfonction']+rule.total
                if rule.code == 'RIndFonction':
                    total['rappel_indfonction'] = total['rappel_indfonction']+rule.total
                if rule.code == 'IndPerformance':
                    total['indperformance'] = total['indperformance']+rule.total
                if rule.code == 'RIndPerformance':
                    total['rappel_indperformance'] = total['rappel_indperformance']+rule.total
                if rule.code == 'rappel_prime_respo':
                    total['rappel_responsabilite'] = total['rappel_responsabilite']+rule.total
                if rule.code == 'rappel_IndRep':
                    total['rappel_representation'] = total['rappel_representation']+rule.total
                if rule.code == 'remb':
                    total['remb'] = total['remb']+rule.total
                if rule.code == 'pret_immob':
                    total['pret_immob'] = total['pret_immob']+rule.total
                if rule.code == 'epargne_retraite_v':
                    total['epargne_retraite_v'] = total['epargne_retraite_v']+rule.total
                if rule.code == 'regul_ind_caisse':
                    total['regul_ind_caisse'] = total['regul_ind_caisse']+rule.total
                if rule.code == 'prime_merite':
                    total['prime_merite'] = total['prime_merite']+rule.total
                if rule.code == 'PRIME_ASTREINTE':
                    total['prime_astreinte'] = total['prime_astreinte']+rule.total
                if rule.code == 'prime_caisse':
                    total['prime_caisse'] = total['prime_caisse']+rule.total
                if rule.code == 'prime_except':
                    total['prime_except'] = total['prime_except']+rule.total
                if rule.code == 'stc':
                    total['stc'] = total['stc']+rule.total
                if rule.code == 'reprise_absence_irr':
                    total['reprise_absence_irr'] = total['reprise_absence_irr']+rule.total
                if rule.code == 'repriseconge':
                    total['repriseconge'] = total['repriseconge']+rule.total
                if rule.code == 'congespayes':
                    total['congespayes'] = total['congespayes']+rule.total
                if rule.code == 'rappel_salaire_jrs':
                    total['rappel_salaire_jrs'] = total['rappel_salaire_jrs']+rule.total
                if rule.code == 'frais_dep':
                    total['fraisdeplacement']=total['fraisdeplacement']+rule.total
                if rule.code == 'allocation':
                    total['allocation']=total['allocation']+rule.total
                if rule.code == 'rappel_allocation':
                    total['rappel_allocation']=total['rappel_allocation']+rule.total
                if rule.code == 'ded_pret_immob':
                    total['ded_logement']=total['ded_logement']+rule.total
                if rule.code == 'rcar_rg':
                    total['rcarplafonne']=total['rcarplafonne']+rule.total
                if rule.code == 'rcar_rc':
                    total['rcarcomplement']=total['rcarcomplement']+rule.total
                if rule.code == 'caad':
                    total['caad']=total['caad']+rule.total
                if rule.code == 'partsalsm':
                    total['modep'] = total['modep'] + rule.total
                if rule.code == 'part_sal_sc':
                    total['cnops']=total['cnops']+rule.total
                if rule.code == 'reprise':
                    total['reprise_abs'] = total['reprise_abs'] + rule.total
                if rule.code == 'issaf':
                    total['saham']=total['saham']+rule.total
                if rule.code == u'congpayé2':
                    total['congpaye2']=total['congpaye2']+rule.total
                if rule.code == 'reprcong2':
                    total['reprcong2']=total['reprcong2']+rule.total
                if rule.code == 'reprisesalaire':
                    total['reprise']=total['reprise']+rule.total
                if rule.code == 'rappel':
                    total['rappelsursalaire']=total['rappelsursalaire']+rule.total
                if rule.code == 'prime_respo':
                    total['responsabilite']=total['responsabilite']+rule.total
                if rule.code == ' BrutCNSSma ':
                    total['brutcnssma']=total['brutcnssma']+rule.total
                if rule.code == 'CS':
                    total['cot_sal']=total['cot_sal']+rule.total
                if rule.code == 'TCOMPMA':
                    total['cot_pat']=total['cot_pat']+rule.total
                if rule.code == 'NET':
                    total['net']=total['net']+rule.total
                if rule.code == 'C_IMP':
                    total['net_imp']=total['net_imp']+rule.total
                if rule.code == 'C_IMPR':
                    total['net_impr']=total['net_impr']+rule.total
                if rule.code == 'TOTAL':
                    total['cout']=total['cout']+rule.total
                if rule.code == 'BASE':
                    total['base']=total['base']+rule.total
                if rule.code == 'ANCIENNETE':
                    total['anciennete']=total['anciennete']+rule.total
                if rule.code == 'IndRep':
                    total['representation']=total['representation']+rule.total
                if rule.code == 'IndemniteTransport':
                    total['transport']=total['transport']+rule.total
                if rule.code == 'regul_ind_transport_v':
                    total['regul_trans']=total['regul_trans']+rule.total
                if rule.code == 'regul_ind_represent_v':
                    total['regul_rep'] = total['regul_rep'] + rule.total
                if rule.code == 'prime_caisse':
                    total['caisse']=total['caisse']+rule.total
                if rule.code == 'BRUTNIMPOSABLE':
                    total['totbrut']=total['totbrut']+rule.total
                if rule.code == 'CNSS ma':
                    total['cotcnss']=total['cotcnss']+rule.total
                if rule.code == 'C1ma':
                    total['cotamo']=total['cotamo']+rule.total
                if rule.code == 'FPRO':
                    total['fpro']=total['fpro']+rule.total
                if rule.code == 'SALC':
                    total['totcot']=total['totcot']+rule.total
                if rule.code == 'IRPP':
                    total['ir']=total['ir']+rule.total
                if rule.code == 'arrondiEnCours':
                    total['arrencours']=total['arrencours']+rule.total
                if rule.code == 'arr_prec':
                    total['arrprec']=total['arrprec']+rule.total
                if rule.code == 'NbrPerso':
                    total['nbrpersoch']=total['nbrpersoch']+rule.total
                if rule.code == 'DEDPERSO':
                    total['dedperso']=total['dedperso']+rule.total
                if rule.code == 'notes_frais':
                    total['notes_frais']=total['notes_frais']+rule.total
                if rule.code == 'regul_jr_ir':
                    total['presence_jrs']=total['presence_jrs']+rule.total
                if rule.code == 'iccp':
                    total['iccp']=total['iccp']+rule.total
                if rule.code == 'PRIME_NAISSANCE':
                    total['prime_naissance']=total['prime_naissance']+rule.total
                if rule.code == 'reprisesalaire':
                    total['reprisesalaire']=total['reprisesalaire']+rule.total
                if rule.code == 'recore_sal':
                    total['recore_sal']=total['recore_sal']+rule.total
                if rule.code == 'prime_scolarite':
                    total['prime_scolarite'] = total['prime_scolarite']+rule.total
                if rule.code == 'prime_achoura_impo':
                    total['prime_achoura_impo'] = total['prime_achoura_impo']+rule.total
                if rule.code == 'prime_achoura':
                    total['prime_achoura'] = total['prime_achoura']+rule.total
            total['nbr_jours']+=i.nbr_jours
            total['jours_acquis']+=i.jours_acquis

        total['brutcnssma'] = locale.format("%.2f", total['brutcnssma'], grouping=True)
        total['recore_sal'] = locale.format("%.2f", total['recore_sal'], grouping=True)
        total['avance'] = locale.format("%.2f", total['avance'], grouping=True)
        total['prime_adha'] = locale.format("%.2f", total['prime_adha'], grouping=True)
        total['indlogement'] = locale.format("%.2f", total['indlogement'], grouping=True)
        total['rappel_indlogement'] = locale.format("%.2f", total['rappel_indlogement'], grouping=True)
        total['indfonction'] = locale.format("%.2f", total['indfonction'], grouping=True)
        total['rappel_indfonction'] = locale.format("%.2f", total['rappel_indfonction'], grouping=True)
        total['indperformance'] = locale.format("%.2f", total['indperformance'], grouping=True)
        total['rappel_indperformance'] = locale.format("%.2f", total['rappel_indperformance'], grouping=True)
        total['rappel_responsabilite'] = locale.format("%.2f", total['rappel_responsabilite'], grouping=True)
        total['rappel_representation'] = locale.format("%.2f", total['rappel_representation'], grouping=True)
        total['remb'] = locale.format("%.2f", total['remb'], grouping=True)
        total['pret_immob'] = locale.format("%.2f", total['pret_immob'], grouping=True)
        total['epargne_retraite_v'] = locale.format("%.2f", total['epargne_retraite_v'], grouping=True)
        total['regul_ind_caisse'] = locale.format("%.2f", total['regul_ind_caisse'], grouping=True)
        total['prime_merite'] = locale.format("%.2f", total['prime_merite'], grouping=True)
        total['prime_except'] = locale.format("%.2f", total['prime_except'], grouping=True)
        total['prime_caisse'] = locale.format("%.2f", total['prime_caisse'], grouping=True)
        total['stc'] = locale.format("%.2f", total['stc'], grouping=True)
        total['reprise_absence_irr'] = locale.format("%.2f", total['reprise_absence_irr'], grouping=True)
        total['repriseconge'] = locale.format("%.2f", total['repriseconge'], grouping=True)
        total['congespayes'] = locale.format("%.2f", total['congespayes'], grouping=True)
        total['rappel_salaire_jrs'] = locale.format("%.2f", total['rappel_salaire_jrs'], grouping=True)
        total['presence']=locale.format("%.2f", total['presence'], grouping=True)
        total['presence_jrs']=locale.format("%.2f", total['presence_jrs'], grouping=True)

        total['net_imp']=locale.format("%.2f", total['net_imp'], grouping=True)
        total['net']=locale.format("%.2f", total['net'], grouping=True)
        total['cot_pat']=locale.format("%.2f", total['cot_pat'], grouping=True)

        total['cot_sal']=locale.format("%.2f", total['cot_sal'], grouping=True)
        total['cout']=locale.format("%.2f", total['cout'], grouping=True)
        total['base']=locale.format("%.2f", total['base'], grouping=True)

        total['anciennete']=locale.format("%.2f", total['anciennete'], grouping=True)
        total['representation']=locale.format("%.2f", total['representation'], grouping=True)
        total['transport']=locale.format("%.2f", total['transport'], grouping=True)
        total['regul_trans']=locale.format("%.2f", total['regul_trans'], grouping=True)
        total['caisse']=locale.format("%.2f", total['caisse'], grouping=True)
        total['congpaye2']=locale.format("%.2f", total['congpaye2'], grouping=True)
        total['reprcong2']=locale.format("%.2f", total['reprcong2'], grouping=True)
        total['reprise']=locale.format("%.2f", total['reprise'], grouping=True)
        total['reprisesalaire']=locale.format("%.2f", total['reprisesalaire'], grouping=True)
        total['rappelsursalaire']=locale.format("%.2f", total['rappelsursalaire'], grouping=True)
        total['rcarcomplement']=locale.format("%.2f", total['rcarcomplement'], grouping=True)
        total['rcarplafonne']=locale.format("%.2f", total['rcarplafonne'], grouping=True)
        total['saham']=locale.format("%.2f", total['saham'], grouping=True)
        total['caad']=locale.format("%.2f", total['caad'], grouping=True)
        total['modep']=locale.format("%.2f", total['modep'], grouping=True)
        total['conps']=locale.format("%.2f", total['cnops'], grouping=True)
        total['ded_logement']=locale.format("%.2f", total['ded_logement'], grouping=True)
        total['responsabilite']=locale.format("%.2f", total['responsabilite'], grouping=True)
        total['fraisdeplacement']=locale.format("%.2f", total['fraisdeplacement'], grouping=True)
        total['notes_frais']=locale.format("%.2f", total['notes_frais'], grouping=True)

        total['totbrut']=locale.format("%.2f", total['totbrut'], grouping=True)
        total['fpro']=locale.format("%.2f", total['fpro'], grouping=True)
        total['totcot']=locale.format("%.2f", total['totcot'], grouping=True)
        total['ir']=locale.format("%.2f", total['ir'], grouping=True)
        total['arrencours']=locale.format("%.2f", total['arrencours'], grouping=True)
        total['arrprec']=locale.format("%.2f", total['arrprec'], grouping=True)
        total['dedperso']=locale.format("%.2f", total['dedperso'], grouping=True)
        total['nbrpersoch']=locale.format("%.2f", total['nbrpersoch'], grouping=True)
        total['allocation']=locale.format("%.2f", total['allocation'], grouping=True)
        total['rappel_allocation']=locale.format("%.2f", total['rappel_allocation'], grouping=True)
        total['iccp']=locale.format("%.2f", total['iccp'], grouping=True)
        total['prime_naissance']=locale.format("%.2f", total['prime_naissance'], grouping=True)
        total['prime_astreinte']=locale.format("%.2f", total['prime_astreinte'], grouping=True)
        total['prime_achoura'] = locale.format("%.2f", total['prime_achoura'], grouping=True)
        total['prime_achoura_impo'] = locale.format("%.2f", total['prime_achoura_impo'], grouping=True)
        total['prime_scolarite'] = locale.format("%.2f", total['prime_scolarite'], grouping=True)
        total['reprise_abs'] = locale.format("%.2f", total['reprise_abs'], grouping=True)
        return total

    def get_number_bp(self):
        return locale.format("%.2f",len(self.get_bulletin_ids(self.periode.id)), grouping=True)

    @api.multi
    def get_line_ids(self):
        list=self.get_bulletin_done_ids()
        if list:
            return list[0][0].line_ids
        else:
            return False


    '''@api.multi
    def get_bulletin_done_ids_test(self):
        test=self.get_bulletin_done_ids()
        for i in test:
            for t in i:
                for j in t.line_ids:
                    print j.name
                    print j.total'''



    @api.multi
    def export_livre_paie(self):
        #ids = self.get_bulletin_done_ids()
        name = "Livre de paie du %s au %s "\
               % (self.start_date,
                  self.end_date)
        self.get_bulletin_done_ids()

        return {
            'type': 'ir.actions.report.xml',
            'report_name': 'paie.report_document_livre_paie',
            'datas': {
                'model': 'hr.payslip.wizard',
                #'ids': list_ids,
                },
            'name': name,
            }


    @api.multi
    def export_livre_paie_excel(self):
        ids = self.get_bulletin_ids(self.periode.id)
        i = 1
        j = 1
        buf = StringIO()
        #############formatage des cellules###############################"
        ft = Font(name='Calibri', size=15, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
        ###############################definition de fichier excel##############
        wb = Workbook()
        ws = wb.active
        ws.title = unicode("Livre de paie")
        ####################################remplir la feuille excel avec les données########################################
        # txt1 = unicode('Livre de paie', "utf8")
        # title1 = txt1+' '+str(self.periode.name)
        # ce = ws.cell(column=3, row=1, value=title1)
        # ce.font = ft
        ft = Font(name='Calibri',size=13,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color= '000076')

        borders = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        #ce.font = ft
        #my_style = Style(border=borders)
        #ce.border=borders
        ce = ws.cell(column=1, row=i, value='Matricule')
        ce.font = ft
        ce = ws.cell(column=2, row=i, value='Nom')
        ce.font = ft
        ce = ws.cell(column=3, row=i, value='Prénom')
        ce.font = ft
        ce = ws.cell(column=4, row=i, value='''Date d'entrée''')
        ce.font = ft
        ce = ws.cell(column=5, row=i, value='Fonction')
        ce.font = ft
        ce = ws.cell(column=6, row=i, value='Salaire de Base')
        ce.font = ft
        ce = ws.cell(column=7, row=i, value='Congés payés')
        ce.font = ft
        ce = ws.cell(column=8, row=i, value='Reprise/Congés')
        ce.font = ft
        ce = ws.cell(column=9, row=i, value='Congés payés–')
        ce.font = ft
        ce = ws.cell(column=10, row=i, value='Reprise/congé–')
        ce.font = ft
        ce = ws.cell(column=11, row=i, value='Jours ICCP')
        ce.font = ft
        ce = ws.cell(column=12, row=i, value='Rappel sur Salaire')
        ce.font = ft
        ce = ws.cell(column=13, row=i, value='Rappel/Salaire Jrs')
        ce.font = ft
        ce = ws.cell(column=14, row=i, value='Reprise/Salaire')
        ce.font = ft
        ce = ws.cell(column=15, row=i, value='Reprise sur salaire')
        ce.font = ft
        ce = ws.cell(column=16, row=i, value='''Absence STC''')
        ce.font = ft
        ce = ws.cell(column=17, row=i, value='''Reprise absence irrégulière''')
        ce.font = ft
        ce = ws.cell(column=18, row=i, value='Indémnité de logement')
        ce.font = ft
        ce = ws.cell(column=19, row=i, value='Rappel Ind de logement')
        ce.font = ft
        ce = ws.cell(column=20, row=i, value='Indémnité de fonction')
        ce.font = ft
        ce = ws.cell(column=21, row=i, value='Rappel Indémnité de fonction')
        ce.font = ft
        ce = ws.cell(column=22, row=i, value='Indemnité de performance')
        ce.font = ft
        ce = ws.cell(column=23, row=i, value='Rappel Indemnité de performance')
        ce.font = ft
        ce = ws.cell(column=24, row=i, value='Prime ancienneté')
        ce.font = ft
        ce = ws.cell(column=25, row=i, value='Prime exceptionnelle')
        ce.font = ft
        ce = ws.cell(column=26, row=i, value='Prime de merite')
        ce.font = ft
        ce = ws.cell(column=27, row=i, value='Prime de naissance')
        ce.font = ft
        ce = ws.cell(column=28, row=i, value="Prime d'astreinte")
        ce.font = ft
        ce = ws.cell(column=29, row=i, value='Prime de responsabilité')
        ce.font = ft
        ce = ws.cell(column=30, row=i, value='Rappel prime responsalibité')
        ce.font = ft
        ce = ws.cell(column=31, row=i, value='Prime de scolarité')
        ce.font = ft
        ce = ws.cell(column=32, row=i, value='Prime adha')
        ce.font = ft
        ce = ws.cell(column=33, row=i, value='Prime achoura impo')
        ce.font = ft
        ce = ws.cell(column=34, row=i, value='Prime achoura')
        ce.font = ft
        ce = ws.cell(column=35, row=i, value='Indemnité de Représentation')
        ce.font = ft
        ce = ws.cell(column=36, row=i, value='Rappel Ind de Représentation')
        ce.font = ft
        ce = ws.cell(column=37, row=i, value='Régul Ind de Représentation')
        ce.font = ft
        ce = ws.cell(column=38, row=i, value='Indemnité de Transport')
        ce.font = ft
        ce = ws.cell(column=39, row=i, value='Régul indemnité de Transport')
        ce.font = ft
        ce = ws.cell(column=40, row=i, value='Indemnité de caisse')
        ce.font = ft
        ce = ws.cell(column=41, row=i, value='Régul indemnité de caisse')
        ce.font = ft
        ce = ws.cell(column=42, row=i, value='Total Brut')
        ce.font = ft
        ce = ws.cell(column=43, row=i, value='RCAR RG PLAFONNE')
        ce.font = ft
        ce = ws.cell(column=44, row=i, value='RCAR RG COMPLEMENT')
        ce.font = ft
        ce = ws.cell(column=45, row=i, value='Secteur Commun CNOPS')
        ce.font = ft
        ce = ws.cell(column=46, row=i, value='Secteur Mutualise MODE')
        ce.font = ft
        ce = ws.cell(column=47, row=i, value='CAAD MODEP')
        ce.font = ft
        ce = ws.cell(column=48, row=i, value='SAHAM Assitance')
        ce.font = ft
        ce = ws.cell(column=49, row=i, value='Retraite complémentaire')
        ce.font = ft
        ce = ws.cell(column=50, row=i, value='Cotisation RECORE')
        ce.font = ft
        ce = ws.cell(column=51, row=i, value='Frai Professionnels')
        ce.font = ft
        ce = ws.cell(column=52, row=i, value='Cotisation salariales')
        ce.font = ft
        ce = ws.cell(column=53, row=i, value='HT Prêt Habitat principal')
        ce.font = ft
        ce = ws.cell(column=54, row=i, value='Déduction logement')
        ce.font = ft
        ce = ws.cell(column=55, row=i, value='Net imposable')
        ce.font = ft
        ce = ws.cell(column=56, row=i, value='Prélévement IR')
        ce.font = ft
        ce = ws.cell(column=57, row=i, value='Remb All.Familiale')
        ce.font = ft
        ce = ws.cell(column=58, row=i, value='Rappel All.Familiale')
        ce.font = ft
        ce = ws.cell(column=59, row=i, value='Remb Frais Déplacement')
        ce.font = ft
        ce = ws.cell(column=60, row=i, value='Remb Notes de frais')
        ce.font = ft
        ce = ws.cell(column=61, row=i, value='Avance sur salaire')
        ce.font = ft
        ce = ws.cell(column=62, row=i, value='Acompte régulier Remb Ecart Retraite Compl')
        ce.font = ft
        ce = ws.cell(column=63, row=i, value='Arrondi du mois précédent')
        ce.font = ft
        ce = ws.cell(column=64, row=i, value='Arrondi du mois en cours')
        ce.font = ft
        ce = ws.cell(column=65, row=i, value='Nbre personnes à Charges')
        ce.font = ft
        ce = ws.cell(column=66, row=i, value='Déduction Personnes à charges')
        ce.font = ft
        ce = ws.cell(column=67, row=i, value='Présence Jrs du mois')
        ce.font = ft
        ce = ws.cell(column=68, row=i, value='Brut soumis à CNSS')
        ce.font = ft
        ce = ws.cell(column=69, row=i, value='Cotisations Patronales')
        ce.font = ft
        ce = ws.cell(column=70, row=i, value='Net à payer')
        ce.font = ft
        ce = ws.cell(column=71, row=i, value='Total des hres travaillées')
        ce.font = ft
        ce = ws.cell(column=72, row=i, value='Congés acquis dans le mois')
        ce.font = ft
        ce = ws.cell(column=73, row=i, value='Congés pris dans le mois')
        ce.font = ft
        ce = ws.cell(column=74, row=i, value='Coût total')
        ce.font = ft
        ce = ws.cell(column=75, row=i, value='Nombre de BP')
        ce.font = ft
        jours_acquis_tot = 0
        jours_pris_tot = 0

        i=i+1

        for bp in ids:
            ce = ws.cell(column=j, row=i, value=bp.employee_id.matricule)
            j += 1
            ce = ws.cell(column=j, row=i, value=bp.employee_id.nom)
            j += 1
            ce = ws.cell(column=j, row=i, value=bp.employee_id.prenom)
            j += 1
            ce = ws.cell(column=j, row=i, value=bp.employee_id.date_debut)
            j += 1
            ce = ws.cell(column=j, row=i, value=bp.employee_id.job_id.name)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "BASE":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == u"congpayé2":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "reprcong2":
                    ce = ws.cell(column=j, row=i, value=-rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "congespayes":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "repriseconge":
                    ce = ws.cell(column=j, row=i, value=-rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "iccp":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "rappel":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "rappel_salaire_jrs":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "reprisesalaire":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "reprise_absence_irr":
                    ce = ws.cell(column=j, row=i, value=-rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "stc":
                    ce = ws.cell(column=j, row=i, value=-rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "reprise":
                    ce = ws.cell(column=j, row=i, value=-rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "IndLogement":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "RIndLogement":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "IndFonction":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "RIndFonction":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "IndPerformance":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "RIndPerformance":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "ANCIENNETE":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_except":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_merite":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "PRIME_NAISSANCE":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "PRIME_ASTREINTE":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_respo":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "rappel_prime_respo":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_scolarite":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_adha":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_achoura_impo":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_achoura":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "IndRep":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "rappel_IndRep":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "regul_ind_represent_v":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "IndemniteTransport":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "regul_ind_transport_v":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "prime_caisse":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "regul_ind_caisse":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "BRUTNIMPOSABLE":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "rcar_rg":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "rcar_rc":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "part_sal_sc":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "partsalsm":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "caad":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "issaf":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "epargne_retraite_v":
                    ce = ws.cell(column=j, row=i, value=self.get_retraite_comp(bp.employee_id.id))
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "recore_sal":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "FPRO":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "CS":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "pret_immob":
                    ce = ws.cell(column=j, row=i, value=self.get_pret(bp.employee_id.id)['paid_amount'])
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "ded_pret_immob":
                    ce = ws.cell(column=j, row=i, value=self.get_deduction_logement(bp.employee_id.id))
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "C_IMPR":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "IRPP":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "allocation":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "rappel_allocation":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "frais_dep":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "notes_frais":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "avance":
                    ce = ws.cell(column=j, row=i, value=-rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "remb":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "arr_prec":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "arrondiEnCours":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "NbrPerso":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "DEDPERSO":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "regul_jr_ir":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == " BrutCNSSma ":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "TCOMPMA":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "NET":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "HoraireMaroc":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            j += 1
            ce = ws.cell(column=j, row=i, value=bp.jours_acquis)
            jours_acquis_tot = jours_acquis_tot + bp.jours_acquis
            j += 1
            ce = ws.cell(column=j, row=i, value=self.get_conge_pris(bp.employee_id.id))
            jours_pris_tot = jours_pris_tot + self.get_conge_pris(bp.employee_id.id)
            j += 1
            for rule in bp.details_by_salary_rule_category:
                if rule.code == "TOTAL":
                    ce = ws.cell(column=j, row=i, value=rule.total)
            i += 1
            j = 1
        ftt = Font(name='Calibri', size=13, bold=True, italic=False, vertAlign=None, underline='none', strike=False)

        ce.font = ftt
        ce = ws.cell(column=6, row=i, value=self.get_sum()['base'])
        ce.font = ftt
        ce = ws.cell(column=7, row=i, value=self.get_sum()['congpaye2'])
        ce.font = ftt
        if self.get_sum()['reprcong2'] == '0,00':
            ce = ws.cell(column=8, row=i, value='0,00')
        else:
            ce = ws.cell(column=8, row=i, value='-'+self.get_sum()['reprcong2'])
        ce.font = ftt
        ce = ws.cell(column=9, row=i, value=self.get_sum()['congespayes'])
        ce.font = ftt
        if self.get_sum()['repriseconge'] == '0,00':
            ce = ws.cell(column=10, row=i, value='0,00')
        else:
            ce = ws.cell(column=10, row=i, value='-'+self.get_sum()['repriseconge'])
        ce.font = ftt
        ce = ws.cell(column=11, row=i, value=self.get_sum()['iccp'])
        ce.font = ftt
        ce = ws.cell(column=12, row=i, value=self.get_sum()['rappelsursalaire'])
        ce.font = ftt
        ce = ws.cell(column=13, row=i, value=self.get_sum()['rappel_salaire_jrs'])
        ce.font = ftt
        ce = ws.cell(column=14, row=i, value=self.get_sum()['reprise'])
        ce.font = ftt
        if self.get_sum()['reprise_absence_irr'] == '0,00':
            ce = ws.cell(column=15, row=i, value='0,00')
        else:
            ce = ws.cell(column=15, row=i, value='-'+self.get_sum()['reprise_absence_irr'])
        ce.font = ftt
        if self.get_sum()['stc'] == '0,00':
            ce = ws.cell(column=16, row=i, value='0,00')
        else:
            ce = ws.cell(column=16, row=i, value='-'+self.get_sum()['stc'])
        ce.font = ftt
        if self.get_sum()['reprise_abs'] == '0,00':
            ce = ws.cell(column=17, row=i, value='0,00')
        else:
            ce = ws.cell(column=17, row=i, value='-'+self.get_sum()['reprise_abs'])
        ce.font = ftt
        ce = ws.cell(column=18, row=i, value=self.get_sum()['indlogement'])
        ce.font = ftt
        ce = ws.cell(column=19, row=i, value=self.get_sum()['rappel_indlogement'])
        ce.font = ftt
        ce = ws.cell(column=20, row=i, value=self.get_sum()['indfonction'])
        ce.font = ftt
        ce = ws.cell(column=21, row=i, value=self.get_sum()['rappel_indfonction'])
        ce.font = ftt
        ce = ws.cell(column=22, row=i, value=self.get_sum()['indperformance'])
        ce.font = ftt
        ce = ws.cell(column=23, row=i, value=self.get_sum()['rappel_indperformance'])
        ce.font = ftt
        ce = ws.cell(column=24, row=i, value=self.get_sum()['anciennete'])
        ce.font = ftt
        ce = ws.cell(column=25, row=i, value=self.get_sum()['prime_except'])
        ce.font = ftt
        ce = ws.cell(column=26, row=i, value=self.get_sum()['prime_merite'])
        ce.font = ftt
        ce = ws.cell(column=27, row=i, value=self.get_sum()['prime_naissance'])
        ce.font = ftt
        ce = ws.cell(column=28, row=i, value=self.get_sum()['prime_astreinte'])
        ce.font = ftt
        ce = ws.cell(column=29, row=i, value=self.get_sum()['responsabilite'])
        ce.font = ftt
        ce = ws.cell(column=30, row=i, value=self.get_sum()['rappel_responsabilite'])
        ce.font = ftt
        ce = ws.cell(column=31, row=i, value=self.get_sum()['prime_scolarite'])
        ce.font = ftt
        ce = ws.cell(column=32, row=i, value=self.get_sum()['prime_adha'])
        ce.font = ftt
        ce = ws.cell(column=33, row=i, value=self.get_sum()['prime_achoura_impo'])
        ce.font = ftt
        ce = ws.cell(column=34, row=i, value=self.get_sum()['prime_achoura'])
        ce.font = ftt
        ce = ws.cell(column=35, row=i, value=self.get_sum()['representation'])
        ce.font = ftt
        ce = ws.cell(column=36, row=i, value=self.get_sum()['rappel_representation'])
        ce.font = ftt
        ce = ws.cell(column=37, row=i, value=self.get_sum()['regul_rep'])
        ce.font = ftt
        ce = ws.cell(column=38, row=i, value=self.get_sum()['transport'])
        ce.font = ftt
        ce = ws.cell(column=39, row=i, value=self.get_sum()['regul_trans'])
        ce.font = ftt
        ce = ws.cell(column=40, row=i, value=self.get_sum()['prime_caisse'])
        ce.font = ftt
        ce = ws.cell(column=41, row=i, value=self.get_sum()['regul_ind_caisse'])
        ce.font = ftt
        ce = ws.cell(column=42, row=i, value=self.get_sum()['totbrut'])
        ce.font = ftt
        ce = ws.cell(column=43, row=i, value=self.get_sum()['rcarplafonne'])
        ce.font = ftt
        ce = ws.cell(column=44, row=i, value=self.get_sum()['rcarcomplement'])
        ce.font = ftt
        ce = ws.cell(column=45, row=i, value=self.get_sum()['cnops'])
        ce.font = ftt
        ce = ws.cell(column=46, row=i, value=self.get_sum()['modep'])
        ce.font = ftt
        ce = ws.cell(column=47, row=i, value=self.get_sum()['caad'])
        ce.font = ftt
        ce = ws.cell(column=48, row=i, value=self.get_sum()['saham'])
        ce.font = ftt
        ce = ws.cell(column=49, row=i, value=self.get_total_retraite_comp())
        ce.font = ftt
        ce = ws.cell(column=50, row=i, value=self.get_sum()['recore_sal'])
        ce.font = ftt
        ce = ws.cell(column=51, row=i, value=self.get_sum()['fpro'])
        ce.font = ftt
        ce = ws.cell(column=52, row=i, value=self.get_sum()['cot_sal'])
        ce.font = ftt
        ce = ws.cell(column=53, row=i, value=self.get_total_pret()['total_paid_amount'])
        ce.font = ftt
        ce = ws.cell(column=54, row=i, value=self.get_total_deduction_logement())
        ce.font = ftt
        ce = ws.cell(column=55, row=i, value=self.get_sum()['net_impr'])
        ce.font = ftt
        ce = ws.cell(column=56, row=i, value=self.get_sum()['ir'])
        ce.font = ftt
        ce = ws.cell(column=57, row=i, value=self.get_sum()['allocation'])
        ce.font = ftt
        ce = ws.cell(column=58, row=i, value=self.get_sum()['rappel_allocation'])
        ce.font = ftt
        ce = ws.cell(column=59, row=i, value=self.get_sum()['fraisdeplacement'])
        ce.font = ftt
        ce = ws.cell(column=60, row=i, value=self.get_sum()['notes_frais'])
        ce.font = ftt
        if self.get_sum()['avance'] == '0,00':
            ce = ws.cell(column=61, row=i, value='0,00')
        else:
            ce = ws.cell(column=61, row=i, value='-'+self.get_sum()['avance'])
        ce.font = ftt
        ce = ws.cell(column=62, row=i, value=self.get_sum()['remb'])
        ce.font = ftt
        ce = ws.cell(column=63, row=i, value=self.get_sum()['arrprec'])
        ce.font = ftt
        ce = ws.cell(column=64, row=i, value=self.get_sum()['arrencours'])
        ce.font = ftt
        ce = ws.cell(column=65, row=i, value=self.get_sum()['nbrpersoch'])
        ce.font = ftt
        ce = ws.cell(column=66, row=i, value=self.get_sum()['dedperso'])
        ce.font = ftt
        ce = ws.cell(column=67, row=i, value=self.get_sum()['presence_jrs'])
        ce.font = ftt
        ce = ws.cell(column=68, row=i, value=self.get_sum()['brutcnssma'])
        ce.font = ftt
        ce = ws.cell(column=69, row=i, value=self.get_sum()['cot_pat'])
        ce.font = ftt
        ce = ws.cell(column=70, row=i, value=self.get_sum()['net'])
        ce.font = ftt
        ce = ws.cell(column=71, row=i, value=self.get_sum()['presence'])
        ce.font = ftt
        ce = ws.cell(column=72, row=i, value=self.get_sum()['jours_acquis'])
        ce.font = ftt
        ce = ws.cell(column=73, row=i, value=self.get_total_conge_pris())
        ce.font = ftt
        ce = ws.cell(column=74, row=i, value=self.get_sum()['cout'])
        ce.font = ftt
        ce = ws.cell(column=75, row=i, value=self.get_number_bp())
        ce.font = ftt
        wb.save(buf)
        fichier = "Livre de paie"+str(self.periode.name)+".xlsx"
        out=base64.encodestring(buf.getvalue())
        buf.close()
        vals = {'data':out,'name_file':fichier}
        wizard_id = self.pool.get('livre.paie.wizard').create(self._cr, self._uid, vals, context=self._context)
        return {
            'name':"Rapport Excel test",
            'view_mode': 'form',
            'view_id': False,
            'view_type': 'form',
            'res_model': 'livre.paie.wizard',
            'res_id':wizard_id,
            'type': 'ir.actions.act_window',
            'target': 'new',
            'domain': '[]',
            }


class livreexcel(models.TransientModel):
    _name = 'livre.paie.wizard'

    name_file = fields.Char('Nom fichier')
    data = fields.Binary('Fichier')